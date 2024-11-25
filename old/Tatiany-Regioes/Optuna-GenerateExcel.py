import ast
import json

from dataclasses import dataclass, field
from functools import singledispatchmethod
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, PatternFill, borders
from openpyxl.worksheet.filters import FilterColumn, Filters
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

DATA_PATH = Path(r'D:\Pedro\Faculdade\FEI-Projeto_Dor\results\Optuna\ViT')

@dataclass
class Header:
    name: str
    parent: Optional['Header'] = field(default=None)
    subheaders: List['Header'] = field(default_factory=list)
    col_span: List[int] = field(default_factory=list)
    row: int = field(default=1)

    def add_subheader(self, subheader: 'Header'):
        # Set the subheader parent as this header
        subheader.parent = self
        # Set the row to be the next row from this header
        subheader.row = self.row+1
        # Add to subheaders list
        self.subheaders.append(subheader)

    def get_subheader(self, h: Union[str,'Header']) -> 'Header':
        try: 
            return self.subheaders[self.subheaders.index(h)]
        except ValueError:
            return None
        
    def flatten(self):
        return [self, *self._flatten(self.subheaders)]

    def _flatten(self, headers:List['Header']):
        flat_subheaders = []
        for h in headers:
            flat_subheaders.append(h)
            if len(h.subheaders) > 0:
                flat_subheaders.extend(self._flatten(h.subheaders))
        return flat_subheaders
    
    @singledispatchmethod
    def set_span(self, span):
        return NotImplementedError(f"Can't set span as type '{type(span)}'")
    @set_span.register
    def _(self, span: int):
        self.col_span.append(span)
    @set_span.register
    def _(self, span: list):
        for s in span:
            self.set_span(s)

    # def _next_letter(self, letter:str) -> str:
    #     # Next letter from the alphabet
    #     return chr(ord(letter)+1)

    def __eq__(self, other: object) -> bool:
        # This method was overwriten to allow compairing header with a string and to use the 'in' operator when Header is inside a list
        value = other if isinstance(other, str) else other.name
        return repr(self) == value
    def __repr__(self) -> str:
        return f"{repr(self.parent)+'/' if self.parent is not None else ''}{self.name}"

@dataclass
class Table:
    headers: List['Header'] = field(default_factory=list)
    _row_data_start: int = field(default=1) # Row to start printing data values
    _col_count: int = field(default=1, init=False) # Count to define col spans for each header

    def add_header(self, h, data=None, parent:Optional[Header]=None):
        # Get the existing header or create a new one and add to the table
        if parent is None: # No parent = use the table
            if (hd:=self.get_header(h)) is not None:
                return hd
            hd = self._add_header(h)
        else: # Has a header parent = query it or create a new one
            if (hd:=parent.get_subheader(h)) is not None:
                return hd
            hd = Header(name=h)
            parent.add_subheader(hd)
        if isinstance(data, dict):
            for key, value in data.items():
                self.add_header(key, value, parent=hd)
        return hd
    def get_header(self, h: Header|str) -> Header:
        try: 
            flatten = self.get_headers_flatten()
            return flatten[flatten.index(h)]
        except ValueError:
            return None
    def get_headers_flatten(self) -> List[Header]:
        headers = []
        for h in self.headers:
            headers.extend(h.flatten())
        return headers
    def assign_columns(self):
        for h in self.headers:
            self._assign_col_span(h)
            
    def _assign_col_span(self, header:Header):
        if len(header.subheaders) == 0: # No sub headers = only 1 column span
            col_span = self._col_count
            self._col_count += 1
            header.col_span = [col_span]
            return col_span
        else:
            header.col_span = []
            for sh in header.subheaders:
                span = self._assign_col_span(sh)
                header.set_span(span)
            return header.col_span

    @singledispatchmethod
    def _add_header(self, h):
        return NotImplementedError(f"Can't add type '{type(h)} as a header'")
    @_add_header.register
    def _(self, h: str):
        h = Header(name=h)
        self._add_header(h)
        return h
    @_add_header.register
    def _(self, h: Header):
        self.headers.append(h)
        return h
    
class TableXLS:
    def __init__(self, data_path:str|Path) -> None:
        self.data_path = data_path
        # Create xls data types
        self.wb = Workbook()
        # Create Table
        self.table = Table()
        # Saves where the header row ends
        self.last_header_row = 1
        # Create table header tree structure
        self._create_header()
        # Add data to the worksheet
        self._write_all_data()
        # Adjust column sizes
        self._fit_col_size()
        # Create filters for all columns
        self._add_filters()

    def to_file(self, file_path:str|Path="out.xlsx"):
        # Save to file
        self.wb.save(str(file_path))

    def _add_filters(self):
        ws = self.wb.active
        filters = self.wb.active.auto_filter
        filters.ref = f"A{self.last_header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

    def _create_header(self):
        # Read all files to understand the header structure
        # some headers may be present in some files but not in others
        iterator = self.data_path.rglob('summary.json')
        for file in iterator:
            with file.open('r') as f:
                data = json.load(f)
            
            for key, value in data.items():
                self.table.add_header(key, data=value)
        # Assign columns to the header tree structure
        self.table.assign_columns()
        # Create columns in the xls worksheet
        ws = self.wb.active
        headers = self.table.get_headers_flatten()
        # Save where the headers end and data starts
        self.last_header_row = max([h.row for h in headers])
        # Header Style
        border = borders.Border(left=borders.Side(style='thick'), 
                                right=borders.Side(style='thick'), 
                                top=borders.Side(style='thick'), 
                                bottom=borders.Side(style='thick'))
        for h in headers:
            cell = ws.cell(row=h.row, column=h.col_span[0], value=h.name)
            # Style cell
            cell.alignment = Alignment(horizontal='center', vertical='top')
            # Merge multicell
            if len(h.subheaders) > 0:
                ws.merge_cells(start_row=h.row, start_column=h.col_span[0], end_row=h.row, end_column=h.col_span[-1])
        # Rowwise style
        merge_rows = {}
        for cells in ws.iter_rows(min_row=1, max_row=max([h.row for h in headers])):
            for cell in cells:
                row = cell.row
                bg_color = "A9A9A9" if row%2 else "D3D3D3"
                if cell.value is None and not isinstance(cell, MergedCell):
                    merge_rows[cell.column] = merge_rows.get(cell.column, [cell.row-1]) + [cell.row]
                else:
                    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type = "solid")
                    cell.border = border
        for col, rows in merge_rows.items():
            ws.merge_cells(start_row=rows[0], start_column=col, end_row=rows[-1], end_column=col)
        # Add filters row
        self.last_header_row+=1 # For the filter
        for i in range(1,ws.max_column+1):
            cell = ws.cell(row=self.last_header_row, column=i, value='Filter')
            cell.fill = PatternFill(start_color='B7C9E2', end_color='B7C9E2', fill_type = "solid")
            cell.alignment = Alignment(horizontal='center')

    def _write_all_data(self):
        last_row_index = {}
        iterator = self.data_path.rglob('summary.json')
        for file in iterator:
            with file.open('r') as f:
                data = json.load(f)
            for key, value in data.items():
                self._write_data(key, value, last_row_index)
            
    def _write_data(self, header:str|Header, data:Any, last_row_index:Dict[str,int]):
            if isinstance(data, dict):
                for key, value in data.items():
                    self._write_data(
                        header=f'{header}/{key}',
                        data=value,
                        last_row_index=last_row_index)
            else:
                h = self.table.get_header(header)
                h_col = h.col_span[0]
                # Get last used row useb by data or by the header in this column
                data_row_idx = last_row_index.get(h_col, self.last_header_row)+1
                # Update the last used row for this column
                last_row_index[h_col] = data_row_idx
                # Add the data to the table
                try:
                    self.wb.active.cell(
                        row=data_row_idx,
                        column=h_col,
                        value=ast.literal_eval(data))
                except:
                    self.wb.active.cell(
                        row=data_row_idx,
                        column=h_col,
                        value=data)

    def _fit_col_size(self):
        # Iterate over all columns and adjust their widths
        for column in self.wb.active.columns:
            max_length = 0
            column_letter = column[-1].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            self.wb.active.column_dimensions[column_letter].width = adjusted_width

table_xls = TableXLS(DATA_PATH)
table_xls.to_file(Path() / "summary.xlsx")